import openpyxl

if __name__ == "__main__":
    # src_file = '../data/my_excel.xlsx'
    src_file = '../data/students.xlsx'

    # FileNotFoundError: [Errno 2] No such file or directory: '../data/my_excel.xlsx'
    wb = openpyxl.load_workbook(src_file)
    sheets = wb.worksheets
    # print(type(sheets))
    # print(sheets)

    # 获取表sheet
    sheet_0 = wb.worksheets[0]
    # sheet_1 = wb['sheet']
    print(f'sheet=[{sheet_0.title}]')

    '''
    获取行，遍历
    '''
    rows = sheet_0.rows
    # print(type(rows))
    # print(rows)

    # row = next(rows)
    # print(type(row))  # <generator object Worksheet._cells_by_row at 0x7f8a20016900>
    # print(row)

    for row in rows:
        # print(type(row))
        # print(row)
        for cell in row:
            # 打印一行内所有单元格
            # coordinate 行列编号
            print(f'{cell.coordinate} - {cell.value}', end='\t|\t')
        # 打印换行
        print()


    '''
    获取区域，变量
    '''
    value = sheet_0['A2:C4']
    print(type(value))  # <class 'tuple'>
    print(value)
    # (
    #   (<Cell 'my_sheet_1'.A2>, <Cell 'my_sheet_1'.B2>, <Cell 'my_sheet_1'.C2>),
    #   (<Cell 'my_sheet_1'.A3>, <Cell 'my_sheet_1'.B3>, <Cell 'my_sheet_1'.C3>),
    #   (<Cell 'my_sheet_1'.A4>, <Cell 'my_sheet_1'.B4>, <Cell 'my_sheet_1'.C4>)
    # )
    for cell_tuple in value:
        for cell in cell_tuple:
            print(f'{cell.coordinate} - {cell.value}', end='\t|\t')
        print()

    print('*' * 50)
    '''
    获取指定行列的数据
    sheet[“A”] — 获取A列的数据
    sheet[“A:C”] — 获取A,B,C三列的数据
    sheet[5] — 只获取第5行的数据
    '''
    # 获取A列的数据
    column_a = sheet_0["A"]
    print(type(column_a))  # <class 'tuple'>
    for cell_in_col in column_a:
        print(f'{cell_in_col.coordinate} - {cell_in_col.value}', end='\t|\t')
    print()
    print('*' * 50)

    # 获取A,B,C三列的数据
    columns_values = sheet_0["A:C"]
    print(type(columns_values))  # <class 'tuple'>
    # 一列的数据
    for cells_tup_in_column in columns_values:
        for cell in cells_tup_in_column:
            print(f'{cell.coordinate} - {cell.value}', end='\t|\t')
        print()

    print('*' * 50)

    # 获取第5行的数据
    row_values = sheet_0[1]
    print(type(row_values))  # <class 'tuple'>
    for cell in row_values:
        print(f'{cell.coordinate} - {cell.value}', end='\t|\t')

    print()
    print('*' * 50)

    '''
    行、列遍历器获取
    '''
    # 获取行列数
    max_row = sheet_0.max_row
    max_column = sheet_0.max_column
    # min_row=None, max_row=None, min_col=None, max_col=None, values_only=False
    # 默认可以不穿关键字参数，就会读取全部
    for row in sheet_0.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            # 打印一行内所有单元格
            # coordinate 行列编号
            print(f'{cell.coordinate} - {cell.value}', end='\t|\t')
        # 打印换行
        print()
    '''
    A2 - (2, 1)	|	B2 - (2, 2)	|	C2 - (2, 3)	|
    A3 - (3, 1)	|	B3 - (3, 2)	|	C3 - (3, 3)	|
    A4 - (4, 1)	|	B4 - (4, 2)	|	C4 - (4, 3)	|
    A5 - zhangsan	|	B5 - 18	|	C5 - 2301	|
    '''

    print('*' * 50)
    for column in sheet_0.iter_cols(min_row=2, max_row=max_row, min_col=2, max_col=max_column):
        for cell_in_col in column:
            print(f'{cell_in_col.coordinate} - {cell_in_col.value}', end='\t|\t')
        print()

    '''
    B2 - (2, 2)	|	B3 - (3, 2)	|	B4 - (4, 2)	|	B5 - 18	|
    C2 - (2, 3)	|	C3 - (3, 3)	|	C4 - (4, 3)	|	C5 - 2301	|
    '''


class ExcelReaderUtils:

    @staticmethod
    def print_excel(src):
        load_workbook = openpyxl.load_workbook(src)
        sheet = load_workbook.worksheets[0]
        print(f'存在{sheet.max_row}行数据')
        for one_row in sheet.rows:
            for cell in one_row:
                print(f'{cell.coordinate} : {cell.value}', end='\t|\t')
            print()







