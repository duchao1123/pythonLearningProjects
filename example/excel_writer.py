import openpyxl
from example.excel_reader import ExcelReaderUtils


class ExcelWriterUtils:

    @staticmethod
    def writer_excel(file_name, args: list, sheet_title="sheet1"):
        # 创建一个全新的Excel文件
        wb = openpyxl.Workbook()
        # 创建sheet
        # !!!! index 参数必须传递，否者导致写入数据错误
        sheet = wb.create_sheet(title=sheet_title, index=0)
        # 行写入
        for line in args:
            sheet.append(line)
        # 保存文件
        wb.save(file_name)


if __name__ == "__main__":
    # ExcelWriterUtils.writer_excel("../data/test.xlsx", [["a", "b", "c"], ["a", "b", "c"]])
    # ExcelReaderUtils.print_excel("../data/test.xlsx")

    # 创建一个全新的Excel文件
    wb = openpyxl.Workbook()
    # 创建sheet
    sheet = wb.create_sheet(title='my_sheet_1', index=0)

    # 写入
    title_data_row = ['name', 'age', 'no']
    sheet.append(title_data_row)

    # 单元格写入
    for r_i in range(2, 5):
        for c_i in range(1, len(title_data_row) + 1):
            # ValueError: Cannot convert (1, 1) to Excel
            sheet.cell(row=r_i, column=c_i, value=str((r_i, c_i)))

    # 行写入
    for i in range(5, 6):
        """
        :param row: iterable containing values to append
        :type row: iterable
        """
        sheet.append(("zhangsan", "18", "2301"))

    # 改写
    # sheet_0['A1'].value = '[0, 0]'
    # sheet_0.cell(2, 2).value = '[0, 0]'
    # sheet_0.cell(3, 3, '[0, 0]')
    # # 保存
    # wb.save('../data/my_excel.xlsx')

    # 保存文件
    wb.save('../data/test.xlsx')

























