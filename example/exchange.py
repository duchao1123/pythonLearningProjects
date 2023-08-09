import csv
import openpyxl
import os
from pathlib import Path
from example.excel_reader import ExcelReaderUtils


class FormatFileExchangeUtils:

    @staticmethod
    def csv2excel(*, src_name: str, target_name: str):
        """
        将csv型数据读为excel
        :param src_name: csv文件名
        :param target_name: excel文件名
        """
        FormatFileExchangeUtils._check_file(src_name, target_name)
        if not src_name.endswith(".csv"):
            raise ValueError("源文件不是csv格式，请检查！")

        # 读出csv
        contents: [str] = []
        with open(src_name, 'r') as f:
            lines = f.readlines()
            if len(lines) <= 0:
                raise ValueError("源文件内容为空！")
            contents = lines[:]
            f.close()

        # 写入excel
        wb = openpyxl.Workbook()
        if len(wb.worksheets) <= 0:
            sheet = wb.create_sheet(title='sheet', index=0)
        else:
            sheet = wb.worksheets[0]
        reader = csv.reader(contents)
        for line in reader:
            sheet.append(line)
        wb.save(target_name)

    @staticmethod
    def excel2csv(*, src_name: str, target_name: str):
        """
        将excel型数据读为csv
        :param src_name: excel文件名
        :param target_name: csv文件名
        """
        FormatFileExchangeUtils._check_file(src_name, target_name)
        if not src_name.endswith(".xlsx"):
            raise ValueError("源文件不是excel格式，请检查！")

        # 读取excel
        wb = openpyxl.load_workbook(src_name)
        sheets = wb.worksheets
        sheet_lines_values = {}
        for sheet in sheets:
            lines_values = []
            for row in sheet.iter_rows():
                values = [cell.value for cell in row]
                # values = (lambda args: [cell.value for cell in args])(row)
                lines_values.append(values)
            sheet_lines_values[sheet.title] = lines_values

        # 写入csv
        for key, values in sheet_lines_values.items():
            path = Path(target_name)
            pre = path.stem
            pro = path.suffix
            new_name = ''.join([pre, '-', key, pro])
            with Path(path.parent, new_name).open(mode='w') as f:
                writer = csv.writer(f)
                writer.writerows(values)

    @classmethod
    def _check_file(cls, src_name: str, target_name: str):
        if not os.path.exists(src_name) or not os.path.isfile(src_name):
            raise ValueError("源文件错误，请检查！")
        if os.path.exists(target_name) and os.path.isfile(target_name):
            raise ValueError("结果文件已存在，请检查！")
        if src_name == target_name:
            raise ValueError("源文件与结果文件同名，请检查！")


if __name__ == "__main__":
    # FormatFileExchangeUtils.csv2excel(src_name="../data/students.csv", target_name="../data/students.xlsx")
    ExcelReaderUtils.print_excel("../data/students.xlsx")
    FormatFileExchangeUtils.excel2csv(src_name="../data/students.xlsx", target_name="../data/students_by_exchange.csv")

